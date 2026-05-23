"""
build_codebook.py
Generates data/gaia_codebook.xlsx — a three-tab Excel codebook for the
GAIA Global AI Adoption Index dataset.
Run from the repo root:  python3 scripts/build_codebook.py
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT
import os

OUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'gaia_codebook.xlsx')

# ── Colour palette ──────────────────────────────────────────────────────────────
TEAL_DARK   = '0D7A5F'   # deep teal — header bg
TEAL_MID    = '0D9E76'   # mid teal — accent rows
TEAL_LIGHT  = 'E6F5F0'   # near-white teal — alt row
WHITE       = 'FFFFFF'
TEXT_DARK   = '0E1F2E'
BORDER_COL  = 'B0D4CA'

# ── Style helpers ───────────────────────────────────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def thin_border():
    s = Side(style='thin', color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border():
    s = Side(style='thin', color=BORDER_COL)
    return Border(bottom=s)

def make_hdr_font(size=11, bold=True, color=WHITE):
    return Font(name='Calibri', size=size, bold=bold, color=color)

def make_body_font(size=10, bold=False, color=TEXT_DARK):
    return Font(name='Calibri', size=size, bold=bold, color=color)

def make_link_font():
    return Font(name='Calibri', size=10, color='1565C0', underline='single')

# ── Tab 1 — README ──────────────────────────────────────────────────────────────
def build_readme(wb):
    ws = wb.create_sheet('README')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 72

    def row(r, b_val='', c_val='', b_bold=False, c_bold=False,
            b_size=10, c_size=10, row_height=None,
            fill_hex=None, c_link=False, c_wrap=True):
        bc = ws.cell(r, 2, b_val)
        cc = ws.cell(r, 3, c_val)
        bc.font = Font(name='Calibri', size=b_size, bold=b_bold, color=TEXT_DARK)
        cc.font = make_link_font() if c_link else Font(name='Calibri', size=c_size, bold=c_bold, color=TEXT_DARK)
        cc.alignment = Alignment(wrap_text=c_wrap, vertical='top')
        bc.alignment = Alignment(vertical='top')
        if fill_hex:
            bc.fill = PatternFill('solid', fgColor=fill_hex)
            cc.fill = PatternFill('solid', fgColor=fill_hex)
        if row_height:
            ws.row_dimensions[r].height = row_height
        return r + 1

    r = 2

    # Title block
    ws.merge_cells(f'B{r}:C{r}')
    tc = ws.cell(r, 2, 'GAIA — Global AI Adoption Index')
    tc.font = Font(name='Calibri', size=22, bold=True, color=TEAL_DARK)
    tc.fill = PatternFill('solid', fgColor='F0FAF6')
    ws.row_dimensions[r].height = 32
    r += 1

    ws.merge_cells(f'B{r}:C{r}')
    sc = ws.cell(r, 2, 'Codebook · Version 1.0 · May 2026 · Leila Aghabarari · gaiaindex.org')
    sc.font = Font(name='Calibri', size=11, color='5b7a96')
    sc.fill = PatternFill('solid', fgColor='F0FAF6')
    ws.row_dimensions[r].height = 20
    r += 1

    ws.merge_cells(f'B{r}:C{r}')
    ws.cell(r, 2).fill = PatternFill('solid', fgColor='F0FAF6')
    ws.row_dimensions[r].height = 6
    r += 1

    # DOI line
    ws.merge_cells(f'B{r}:C{r}')
    dc = ws.cell(r, 2, 'https://doi.org/10.5281/zenodo.20320112')
    dc.font = make_link_font()
    dc.hyperlink = 'https://doi.org/10.5281/zenodo.20320112'
    dc.fill = PatternFill('solid', fgColor='F0FAF6')
    ws.row_dimensions[r].height = 18
    r += 1
    r += 1  # blank

    # Section: Description
    def section_hdr(row_num, title):
        ws.merge_cells(f'B{row_num}:C{row_num}')
        c = ws.cell(row_num, 2, title)
        c.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
        c.fill = PatternFill('solid', fgColor=TEAL_DARK)
        c.alignment = Alignment(vertical='center', indent=1)
        ws.row_dimensions[row_num].height = 20
        return row_num + 1

    def kv(row_num, key, val, wrap=True, height=None):
        kc = ws.cell(row_num, 2, key)
        kc.font = Font(name='Calibri', size=10, bold=True, color=TEAL_DARK)
        kc.fill = PatternFill('solid', fgColor='F7FBFA')
        kc.alignment = Alignment(vertical='top', indent=1)
        kc.border = bottom_border()
        vc = ws.cell(row_num, 3, val)
        vc.font = Font(name='Calibri', size=10, color=TEXT_DARK)
        vc.fill = PatternFill('solid', fgColor='F7FBFA')
        vc.alignment = Alignment(wrap_text=wrap, vertical='top', indent=1)
        vc.border = bottom_border()
        if height:
            ws.row_dimensions[row_num].height = height
        return row_num + 1

    r = section_hdr(r, 'About GAIA')
    r = kv(r, 'Description',
        'GAIA (Global AI Adoption Index) is an open dataset measuring artificial intelligence '
        'exposure and adoption across 923 occupations and 138 countries. It combines three '
        'independent data sources — theoretical task exposure, pre-GPT machine learning '
        'suitability, and observed real-world AI tool usage — into a unified, freely accessible '
        'index designed for researchers, policymakers, and development finance practitioners.',
        height=60)
    r = kv(r, 'Website', 'https://gaiaindex.org')
    ws.cell(r-1, 3).hyperlink = 'https://gaiaindex.org'
    ws.cell(r-1, 3).font = make_link_font()
    r = kv(r, 'Version', '1.0')
    r = kv(r, 'Date', 'May 2026')
    r = kv(r, 'Author', 'Leila Aghabarari — International Finance Corporation / World Bank')
    r = kv(r, 'License', 'Creative Commons Attribution 4.0 International (CC BY 4.0). '
        'You are free to share and adapt this data with attribution.')
    r += 1

    r = section_hdr(r, 'How to Cite')
    r = kv(r, 'Dataset',
        'Aghabarari, L. (2026). GAIA — Global AI Adoption Index (Version 1.0). Zenodo. '
        'https://doi.org/10.5281/zenodo.20320112', height=30)
    r = kv(r, 'Companion paper',
        'Aghabarari, L. & Van Doornik, B. (2026). AI Exposure, Credit Markets, and Employment '
        'in Brazil. Working paper.', height=30)
    r += 1

    r = section_hdr(r, 'Data Sources')

    sources = [
        ('Anthropic Economic Index',
         'Handa et al. (2025). Which Economic Tasks are Performed with AI? Evidence from Millions '
         'of Claude Conversations. Hugging Face: Anthropic/EconomicIndex. License: CC BY 4.0. '
         'Reference week: February 5–12, 2026.'),
        ('Eloundou et al. (2024)',
         'Eloundou, T., Manning, S., Mishkin, P., & Rock, D. (2024). GPTs are GPTs: Labor Market '
         'Impact Potential of LLMs. Science 384(6702), 1306–1308. '
         'DOI: 10.1126/science.adj0998. Exposure estimates using GPT-4 and human annotators '
         'across 18,000+ O*NET work tasks, 923 SOC-2018 occupations.'),
        ('Brynjolfsson, Mitchell & Rock (2018)',
         'Brynjolfsson, E., Mitchell, T., & Rock, D. (2018). What Can Machines Learn, and What '
         'Does It Mean for Occupations and the Economy? AEA Papers and Proceedings 108, 43–47. '
         'Replication data: ICPSR 114436. License: CC BY 4.0. '
         'Pre-GPT suitability for machine learning — use as falsification instrument.'),
        ('IMF AI Preparedness Index (AIPI)',
         'International Monetary Fund (2023). Measuring Countries\' AI Preparedness. '
         'IMF DataMapper. Indicator: AI_PI. Sub-indices: digital infrastructure, innovation, '
         'human capital, regulation. 130+ countries.'),
    ]

    for src_name, src_desc in sources:
        r = kv(r, src_name, src_desc, height=50)

    r += 1
    r = section_hdr(r, 'File Contents')
    r = kv(r, 'gaia_occupations.csv', '923 occupations (US SOC 2018). See tab "Occupation Variables".')
    r = kv(r, 'gaia_countries.csv', '138 countries. See tab "Country Variables".')
    r = kv(r, 'gaia_codebook.xlsx', 'This file.')

    # Disclaimer
    r += 1
    ws.merge_cells(f'B{r}:C{r}')
    disc = ws.cell(r, 2,
        'Disclaimer: The views, analysis, and content in this dataset are entirely those of the '
        'author and do not represent the views of the International Finance Corporation, the World '
        'Bank Group, or any affiliated organization.')
    disc.font = Font(name='Calibri', size=9, italic=True, color='5b7a96')
    disc.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 30

    return ws


# ── Tab 2 & 3 — Variable tables ──────────────────────────────────────────────────
COLS = ['Variable Name', 'Full Name', 'Description', 'Data Type',
        'Range / Values', 'Source', 'Notes']
COL_WIDTHS = [22, 26, 52, 12, 22, 30, 32]

OCC_VARS = [
    # (var_name, full_name, description, dtype, range_vals, source, notes)
    ('O*NET-SOC Code', 'Occupation Code',
     'US Standard Occupational Classification (2018) code, 6-digit with decimal.',
     'String', 'e.g. 11-1011.00',
     'BLS / O*NET', 'Primary key. Decimal form: XX-XXXX.XX'),

    ('Title', 'Occupation Title',
     'Full occupation name in English.',
     'String', 'Text',
     'O*NET', ''),

    ('group', 'Major Occupational Group',
     'Broad SOC major group (2-digit prefix). 22 categories covering all occupations.',
     'String', '22 categories',
     'BLS SOC 2018', 'e.g. Management, Healthcare, Computer and Mathematical'),

    ('dv_rating_alpha', 'E1 — AI-rated (no tools)',
     'Share of occupation tasks performable by GPT-4 without any software tools. '
     'Lower bound of AI exposure. AI-annotated version.',
     'Float', '0 – 100',
     'Eloundou et al. Science 2024',
     'Divide by 100 for 0–1 scale. Use dv_rating_beta as main specification.'),

    ('dv_rating_beta', 'E1+E2 — AI-rated (with tools)',
     'Share of occupation tasks performable by GPT-4 augmented with software tools. '
     'MAIN SPECIFICATION for AI exposure in empirical work. AI-annotated version.',
     'Float', '0 – 100',
     'Eloundou et al. Science 2024',
     'Divide by 100 for 0–1. Recommended main independent variable.'),

    ('dv_rating_gamma', 'E1+E2+E3 — AI-rated (broadest)',
     'Broadest exposure tier from AI annotation: tasks performable with full tool integration.',
     'Float', '0 – 100',
     'Eloundou et al. Science 2024',
     'Use as upper-bound robustness check.'),

    ('human_rating_alpha', 'E1 — Human-annotated (no tools)',
     'Human-annotator version of E1. Share of tasks performable by GPT-4 without tools, '
     'rated by human coders rather than the AI model itself.',
     'Float', '0 – 100',
     'Eloundou et al. Science 2024',
     'Use for robustness checks against AI self-rating bias.'),

    ('human_rating_beta', 'E1+E2 — Human-annotated (with tools)',
     'Human-annotator version of E1+E2. Main spec equivalent rated by human coders.',
     'Float', '0 – 100',
     'Eloundou et al. Science 2024',
     'Use for robustness checks. Compare with dv_rating_beta.'),

    ('human_rating_gamma', 'E1+E2+E3 — Human-annotated (broadest)',
     'Broadest exposure tier, human-annotated.',
     'Float', '0 – 100',
     'Eloundou et al. Science 2024',
     ''),

    ('sml_score', 'SML — Suitability for Machine Learning',
     'Pre-GPT measure of how suitable an occupation\'s tasks are for traditional machine learning. '
     'Constructed in 2018 before the generative AI era.',
     'Float', '0 – 100',
     'Brynjolfsson, Mitchell & Rock, AEA 2018. Replication: ICPSR 114436',
     'Falsification instrument: should show no pre-2022 treatment effect in DiD designs.'),

    ('aei_task_success', 'Task Success Rate (Anthropic)',
     'Share of AI tasks within this occupation\'s sessions rated as successfully completed by the model.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index, Feb 2026',
     '177 occupations matched. Missing for unmatched occupations.'),

    ('aei_autonomy_pct', 'AI Autonomy Share (Anthropic)',
     'Share of tasks in which the AI operated autonomously (without significant user correction).',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index, Feb 2026',
     '177 matched occupations.'),

    ('aei_work_pct', 'Work Use Share (Anthropic)',
     'Share of this occupation\'s Claude.ai sessions classified as professional/work use (vs personal or coursework).',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index, Feb 2026',
     '177 matched occupations. Proxy for revealed professional AI adoption.'),

    ('gaia_e', 'GAIA-E Composite Exposure Score',
     'Composite AI exposure index. Computed as the mean of min-max-normalized dv_rating_beta '
     'and min-max-normalized sml_score. Full 923-occupation coverage (sml fallback where '
     'Anthropic data unavailable).',
     'Float', '0.000 – 1.000',
     'GAIA (this dataset)',
     'Recommended headline score for descriptive analysis and cross-occupation comparison.'),
]

CTY_VARS = [
    ('geo_id', 'ISO Country Code (2-letter)',
     'ISO 3166-1 alpha-2 two-letter country code.',
     'String', 'e.g. BR, US, NG',
     'ISO 3166-1', 'Primary key.'),

    ('country_name', 'Country Name',
     'Country name in English.',
     'String', 'Text',
     'ISO / World Bank', ''),

    ('iso3', 'ISO Country Code (3-letter)',
     'ISO 3166-1 alpha-3 three-letter country code.',
     'String', 'e.g. BRA, USA, NGA',
     'ISO 3166-1', 'Used for merging with IMF and Gallup data.'),

    ('income_group', 'World Bank Income Group',
     'World Bank 2025 income classification based on GNI per capita.',
     'String', 'High / Upper-middle / Lower-middle / Low',
     'World Bank 2025', ''),

    ('usage_pct_global', 'Global Claude.ai Traffic Share',
     'This country\'s share of total global Claude.ai traffic during the reference week.',
     'Float', '% (sums to ~100)',
     'Anthropic Economic Index, Feb 5–12 2026',
     'Population-adjusted usage proxy for AI adoption intensity.'),

    ('uc_work', 'Work Use Share',
     'Share of Claude.ai sessions from this country classified as professional/work use.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', ''),

    ('uc_personal', 'Personal Use Share',
     'Share of sessions classified as personal (non-work, non-coursework) use.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', ''),

    ('uc_coursework', 'Coursework / Student Use Share',
     'Share of sessions classified as coursework or student-driven use.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', 'Higher in lower-income countries.'),

    ('uc_none', 'Unclassified Use — None',
     'Share of sessions where use case was not classified (null/none category).',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', 'Typically exclude from use-case breakdowns.'),

    ('uc_not_classified', 'Use Case Not Classified',
     'Share of sessions not classified into any use-case category.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', ''),

    ('collab_directive', 'Directive Collaboration Style',
     'Share of sessions where the user issues direct commands without iteration — '
     'treating the AI as an execution engine.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', 'Higher share indicates more transactional AI use.'),

    ('collab_learning', 'Learning / Exploratory Style',
     'Share of sessions classified as exploratory or learning-oriented — '
     'user asks AI to explain, teach, or elaborate.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', ''),

    ('collab_task_iteration', 'Task Iteration Style',
     'Share of sessions involving back-and-forth task refinement between user and AI.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', ''),

    ('collab_feedback_loop', 'Feedback Loop Style',
     'Share of sessions with revision cycles where user provides feedback on AI output.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', ''),

    ('collab_validation', 'Validation Style',
     'Share of sessions where the user is primarily checking or verifying AI output.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', ''),

    ('collab_none', 'Collaboration Style — None',
     'Sessions with no identifiable collaboration style.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', 'Typically exclude from style breakdowns.'),

    ('collab_not_classified', 'Collaboration Style Not Classified',
     'Sessions not classified into any collaboration style.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', ''),

    ('ts_yes', 'Task Success Rate',
     'Share of tasks rated as successfully completed by the AI.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', 'Primary quality/capability indicator.'),

    ('ts_no', 'Task Failure Rate',
     'Share of tasks rated as not successfully completed.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', 'ts_yes + ts_no + ts_not_classified ≈ 100.'),

    ('ts_not_classified', 'Task Success Not Classified',
     'Share of tasks with indeterminate success classification.',
     'Float', '0 – 100 (%)',
     'Anthropic Economic Index', ''),

    ('aipi_overall', 'IMF AI Preparedness Index (Overall)',
     'IMF AI Preparedness Index — overall composite score across four dimensions.',
     'Float', '0.000 – 1.000',
     'IMF DataMapper 2023. Indicator: AI_PI',
     'Scale: 0 (least prepared) to 1 (most prepared).'),

    ('aipi_digital_infra', 'IMF AIPI — Digital Infrastructure',
     'AIPI sub-index measuring internet connectivity, broadband, and digital infrastructure.',
     'Float', '0.000 – 1.000',
     'IMF 2023', ''),

    ('aipi_innovation', 'IMF AIPI — Innovation Capacity',
     'AIPI sub-index measuring R&D investment, patents, and AI innovation ecosystem.',
     'Float', '0.000 – 1.000',
     'IMF 2023', ''),

    ('aipi_human_capital', 'IMF AIPI — Human Capital',
     'AIPI sub-index measuring STEM education, AI talent, and workforce skills.',
     'Float', '0.000 – 1.000',
     'IMF 2023', ''),

    ('aipi_regulation', 'IMF AIPI — Regulation & Governance',
     'AIPI sub-index measuring AI policy frameworks and regulatory environment.',
     'Float', '0.000 – 1.000',
     'IMF 2023', ''),

    ('gaia_a', 'GAIA-A Adoption Score',
     'Composite adoption index. Mean of min-max-normalized uc_work and min-max-normalized '
     'aipi_overall. Captures both observed behavioral adoption and structural AI readiness.',
     'Float', '0.000 – 1.000',
     'GAIA (this dataset)',
     '128 countries with both components. Recommended for cross-country adoption analysis.'),

    ('gaia_r', 'GAIA-R Readiness Score',
     'AI readiness index. Min-max-normalized aipi_overall. Structural readiness without '
     'behavioral component. Useful for countries with missing Anthropic data.',
     'Float', '0.000 – 1.000',
     'GAIA (this dataset)',
     '133 countries with AIPI data.'),
]


def build_var_tab(wb):
    ws = wb.create_sheet('Variable Descriptions')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    # Subtitle row
    ws.merge_cells('A1:G1')
    sub = ws.cell(1, 1, 'Variable reference for gaia_occupations.csv and gaia_countries.csv   ·   GAIA Global AI Adoption Index  ·  gaiaindex.org')
    sub.font = Font(name='Calibri', size=10, italic=True, color='5b7a96')
    sub.fill = PatternFill('solid', fgColor='F0FAF6')
    sub.alignment = Alignment(vertical='center', indent=1)
    ws.row_dimensions[1].height = 18

    # Header row
    hdr_fill_obj = PatternFill('solid', fgColor=TEAL_DARK)
    for ci, (col_name, col_w) in enumerate(zip(COLS, COL_WIDTHS), 1):
        c = ws.cell(2, ci, col_name)
        c.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
        c.fill = hdr_fill_obj
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = col_w
    ws.row_dimensions[2].height = 22

    highlight_vars = {
        'dv_rating_beta': '0D9E76',
        'gaia_e': TEAL_DARK,
        'gaia_a': TEAL_DARK,
        'gaia_r': TEAL_DARK,
        'aipi_overall': '1565C0',
    }

    def write_section(section_label, var_rows, start_row):
        # Section divider row
        ws.merge_cells(f'A{start_row}:G{start_row}')
        sc = ws.cell(start_row, 1, section_label)
        sc.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
        sc.fill = PatternFill('solid', fgColor=TEAL_MID)
        sc.alignment = Alignment(vertical='center', indent=1)
        ws.row_dimensions[start_row].height = 20
        row_num = start_row + 1

        alt_fill  = PatternFill('solid', fgColor=TEAL_LIGHT)
        white_fill = PatternFill('solid', fgColor=WHITE)

        for ri, var in enumerate(var_rows, 1):
            fill = alt_fill if ri % 2 == 0 else white_fill
            for ci, val in enumerate(var, 1):
                c = ws.cell(row_num, ci, val)
                c.font = Font(name='Calibri', size=10, color=TEXT_DARK)
                c.fill = fill
                c.alignment = Alignment(wrap_text=True, vertical='top', indent=1)
                c.border = thin_border()
                if ci == 1:
                    c.font = Font(name='Calibri', size=10, bold=True, color=TEAL_DARK)
            desc = var[2]
            approx_lines = max(1, len(desc) // 60)
            ws.row_dimensions[row_num].height = max(18, approx_lines * 15 + 6)

            if var[0] in highlight_vars:
                badge_color = highlight_vars[var[0]]
                for ci in range(1, len(COLS) + 1):
                    ws.cell(row_num, ci).fill = PatternFill('solid', fgColor='EAF6F1')
                ws.cell(row_num, 1).font = Font(name='Calibri', size=10, bold=True, color=badge_color)

            row_num += 1
        return row_num

    next_row = write_section('Occupation Variables  (gaia_occupations.csv — 923 occupations)', OCC_VARS, 3)
    next_row += 1  # blank gap between sections
    write_section('Country Variables  (gaia_countries.csv — 138 countries)', CTY_VARS, next_row)

    return ws


# ── Main ────────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]

    build_readme(wb)
    build_var_tab(wb)

    out = os.path.abspath(OUT_PATH)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f'Saved → {out}')
    print(f'  Tabs: {wb.sheetnames}')
    print(f'  Occupation variables: {len(OCC_VARS)}')
    print(f'  Country variables:    {len(CTY_VARS)}')


if __name__ == '__main__':
    main()
